from typing import Optional
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook
from fastapi import File, UploadFile
from sqlalchemy import func
from Settings import setting
from database import get_db
from sqlalchemy.orm import Session
from fastapi import APIRouter, status, UploadFile, Depends
from file_manipulation_product import validate_field_and_add_products
from fastapi import APIRouter, File, UploadFile
import openpyxl
from image_utils import delete_image, upload_image
import models
from models.order_line import OrderLine
import schemas
from fastapi import  Depends, status
from database import get_db
from sqlalchemy.orm import Session
from fastapi import APIRouter, UploadFile, File
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import io
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import utils

router=APIRouter(
    prefix="/product",
    tags=["products"]
)

@router.post("/add" , response_model=schemas.productOut)
async def add_product(product : schemas.productIn ,quantity : int, image: UploadFile = File(None), db: Session = Depends(get_db)):
    try:
        existing_product = db.query(models.Product).filter_by(name=product.name).first()
        if(existing_product):
            existing_product.quantity += quantity
            return schemas.productOut(
            message=" product already exist , quantity updated successfuly",
            status=status.HTTP_200_OK,
        )
        else :
            category = db.query(models.Category).filter_by(name=product.category).first()
            if not category:
                return schemas.productOut(
                    message="category name not found ",
                    status=status.HTTP_400_BAD_REQUEST,
                )
            else :
                try:
                    img_link = None
                    img_public_id = None
                    if image:
                        image_stream = io.BytesIO(await image.read())
                        upload_img_result = await upload_image(image_stream)
                        img_link = upload_img_result["url"]
                        img_public_id = upload_img_result["public_id"]

                    category_id = category.id
                    product_input = models.Product(name = product.name,
                                    category_id = category_id,
                                    description = product.description,
                                    unit_price = product.price,
                                    quantity =quantity ,
                                    image_link =img_link ,)

                    db.add(product_input)
                    db.commit()
                except Exception as e :
                    db.rollback()
                    if img_public_id is not None:
                        delete_image(img_public_id)
                    return schemas.productOut(
                    message="Database error occured",
                    status=status.HTTP_400_BAD_REQUEST,
                )
    except Exception  :
        db.rollback()
        return schemas.productOut(
        message="Database error occured",
        status=status.HTTP_400_BAD_REQUEST,)
    return schemas.productOut(
        message=" product added sucessfly",
        status=status.HTTP_201_CREATED,)


@router.post("/edit" , response_model=schemas.productOut)
async def edit_product(product : schemas.productIn, quantity:int,image: UploadFile = File(None),  db: Session = Depends(get_db)):
    try:
        existing_product = db.query(models.Product).filter_by(id=product.id).first()

        if not existing_product :
            return schemas.productOut(
                message= "product not found",
                status=status.HTTP_400_BAD_REQUEST,
            )

        img_link = None
        img_public_id = None
        if image:
            image_stream = io.BytesIO(await image.read())
            upload_img_result = await upload_image(image_stream)
            img_link = upload_img_result["url"]
            img_public_id = upload_img_result["public_id"]
            existing_product.image_link = img_link
        category = db.query(models.Category).filter_by(name=product.category).first()
        category_id = category.id
        existing_product.name = product.name
        existing_product.category_id = category_id
        existing_product.description = product.description
        existing_product.unit_price = product.price
        existing_product.quantity = quantity
        db.commit()

    except Exception :
        db.rollback()
        if img_public_id is not None:
            delete_image(img_public_id)
        return schemas.productOut(
            message="Database error occured",
            status=status.HTTP_400_BAD_REQUEST,
        )
    return schemas.productOut(
        message=" product edited sucessfly",
        status=status.HTTP_201_CREATED,
    )

@router.post("/delete" , response_model=schemas.productOut)
def delete_product(name : str,  db: Session = Depends(get_db)):
    try:
        existing_product = db.query(models.Product).filter_by(name=name).first()
        if not existing_product :
            return schemas.productOut(
                message= "product name not found",
                status=status.HTTP_400_BAD_REQUEST
            )

        product_id_in_order_line =  db.query(OrderLine).filter_by(product_id=existing_product.id).first()
        if product_id_in_order_line :
            return schemas.productOut(
                message=" We can only delete products which are not used in any order",
                status=status.HTTP_400_BAD_REQUEST,)

        db.delete(existing_product)
        db.commit()

    except Exception as e :
        db.rollback()
        return schemas.productOut(
            message="Database error occured",
            status=status.HTTP_400_BAD_REQUEST,
        )
    return schemas.productOut(
        message=" product deleted successfly",
        status=status.HTTP_201_CREATED,
    )


@router.post("/Upload/" , response_model=schemas.ImportProductsResponse)
async def upload_products(file: UploadFile = File(...),db: Session = Depends(get_db)):

    if not file.filename.endswith('.xlsx'):
        return schemas.ImportProductsResponse(
            message="Invalid file type. Please upload an Excel file.",
            status=status.HTTP_400_BAD_REQUEST
        )

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    sheet = wb.active

    image_loader = SheetImageLoader(sheet)
    image_urls = []

    header = [cell.value for cell in sheet[1]]
    data_list = []

    for row in sheet.iter_rows(min_row=2, max_row=10, min_col=1, max_col=7):
        data = dict(zip(header, [cell.value for cell in row]))
        image_found = False
        for cell in row:
            if image_loader.image_in(cell.coordinate):
                image_found =True
                img = image_loader.get(cell.coordinate)
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)

                try:
                    upload_result = await upload_image(img_byte_arr)
                    file_url = upload_result["url"]
                    data['image_link'] = file_url
                    data_list.append(data)
                    image_urls.append({"url": file_url, "public_id": upload_result["public_id"]})
                except Exception as e:
                    return schemas.ImportProductsResponse(
                        message=f"Image upload failed: {str(e)}",
                        status=status.HTTP_400_BAD_REQUEST
                    )
                break

    if not image_found:
            data['image_link'] = None

    return validate_field_and_add_products(data_list,db)

@router.post("/getall",response_model=schemas.productsOut)
def get_all_products(page_number : int ,page_size:int , name: Optional[str] = None,
    category_name: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_quantity: Optional[int] = None,
    max_quantity: Optional[int] = None,db: Session = Depends(get_db) ):

    product_query = db.query(models.Product)
    if name:
        product_query = product_query.filter(func.lower(
            models.Product.name).contains(func.lower(name)))

    if category_name:
        product_query = product_query.join(models.Category).filter(
            func.lower(models.Category.name) == func.lower(category_name))

    if min_price:
        product_query = product_query.filter(
            models.Product.unit_price >= min_price)

    if max_price:
        product_query = product_query.filter(
            models.Product.unit_price <= max_price)

    if min_quantity:
        product_query = product_query.filter(
            models.Product.quantity >= min_quantity)

    if max_quantity:
        product_query = product_query.filter(
            models.Product.quantity <= max_quantity)

    total_records = db.query(models.Product).count()
    total_pages = utils.div_ceil(total_records,page_size)
    products = product_query.offset((page_number-1) * page_size).limit(page_size).all()

    return schemas.productsOut(
        total_records = total_records,
        total_pages = total_pages,
        page_number = page_number,
        page_size = page_size,
        list = [ schemas.productOut(**product.__dict__) for product in products],
        status=status.HTTP_200_OK )
